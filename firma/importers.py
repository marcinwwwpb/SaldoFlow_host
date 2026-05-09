from __future__ import annotations

from datetime import date, datetime
import hashlib
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from django.contrib.auth import get_user_model
from django.db import transaction

from .models import FakturaKosztowa, ImportDemona, Kontrahent
from finanse.importers import log_demon_event
from paneladmin.utils import audit_log
from finanse.models import DemonLog

OCZEKIWANE_NAGLOWKI_IMPORTU = [
    "Data zakupu",
    "Numer faktury",
    "NIP dostawcy",
    "Netto",
    "VAT",
    "Brutto",
    "Kod kraju",
    "Rodzaj zakupu",
    "Miesiąc JPK",
]


def _normalize_header(value):
    return str(value or "").strip().lower()


def parse_decimal(value):
    if value is None or value == "":
        return Decimal("0.00")
    text = str(value).strip().replace(" ", "").replace(",", ".")
    return Decimal(text).quantize(Decimal("0.01"))


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        raise ValueError("Brak daty")

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Niepoprawna data: {text}")


def parse_month(value, fallback_date):
    if value in (None, ""):
        return fallback_date.month
    text = str(value).strip()
    if text.isdigit():
        month = int(text)
        if 1 <= month <= 12:
            return month
    if "-" in text:
        parts = text.split("-")
        if len(parts) == 2 and parts[1].isdigit():
            month = int(parts[1])
            if 1 <= month <= 12:
                return month
    return fallback_date.month


def map_rodzaj_zakupu(value):
    text = str(value or "").strip().lower()
    srodki_trwale_aliases = {
        "st", "środek trwały", "srodek trwaly", "środki trwałe", "srodki trwale", "fixed", "fixed asset"
    }
    return "ST" if text in srodki_trwale_aliases else "POZ"


def import_koszty_excel_for_user(*, user, fileobj: BinaryIO, source_name: str, source_path: str = "", import_record=None):
    import openpyxl

    workbook = openpyxl.load_workbook(fileobj, data_only=True)
    sheet = workbook.active

    naglowki = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    znormalizowane = [_normalize_header(h) for h in naglowki]
    oczekiwane = [_normalize_header(h) for h in OCZEKIWANE_NAGLOWKI_IMPORTU]

    if znormalizowane != oczekiwane:
        raise ValueError(
            "Nieprawidłowy układ kolumn. Oczekiwane kolumny: " + ", ".join(OCZEKIWANE_NAGLOWKI_IMPORTU)
        )

    dodane = 0
    bledy = []
    with transaction.atomic():
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            try:
                data_zakupu = parse_date(row[0])
                numer_faktury = str(row[1]).strip()
                nip_dostawcy = str(row[2] or "").strip()
                kwota_netto = parse_decimal(row[3])
                kwota_vat = parse_decimal(row[4])
                kwota_brutto = parse_decimal(row[5])
                kod_kraju = str(row[6] or "PL").strip().upper()[:2] or "PL"
                rodzaj_zakupu = map_rodzaj_zakupu(row[7])
                miesiac_jpk = parse_month(row[8], data_zakupu)
                if not numer_faktury:
                    raise ValueError("Brak numeru faktury")

                kontrahent = None
                if nip_dostawcy:
                    kontrahent = Kontrahent.objects.filter(user=user, nip=nip_dostawcy).first()
                    if kontrahent is None:
                        kontrahent = Kontrahent.objects.create(user=user, nazwa=f"Kontrahent {nip_dostawcy}", nip=nip_dostawcy)

                existing = FakturaKosztowa.objects.filter(user=user, numer_faktury=numer_faktury, data_zakupu=data_zakupu, kwota_brutto=kwota_brutto).first()
                if existing:
                    bledy.append(f"Wiersz {idx}: potencjalny duplikat faktury — rekord już istnieje")
                    audit_log(actor=user, module="FIRMA", entity_type="FakturaKosztowa", entity_id=existing.id, action="IMPORT_SKIP", payload={"source_name": source_name, "row": idx, "numer_faktury": numer_faktury})
                    continue
                FakturaKosztowa.objects.create(
                    user=user,
                    kontrahent=kontrahent,
                    kontrahent_nazwa=(kontrahent.nazwa if kontrahent else (nip_dostawcy or "Kontrahent z importu")),
                    numer_faktury=numer_faktury,
                    data_zakupu=data_zakupu,
                    kwota_netto=kwota_netto,
                    kwota_vat=kwota_vat,
                    kwota_brutto=kwota_brutto,
                    nip_dostawcy=nip_dostawcy,
                    kod_kraju=kod_kraju,
                    rodzaj_zakupu=rodzaj_zakupu,
                    miesiac_jpk=miesiac_jpk,
                    kategoria="Import Excel",
                )
                dodane += 1
            except (ValueError, InvalidOperation) as exc:
                bledy.append(f"Wiersz {idx}: {exc}")

    summary = {
        "dodane": dodane,
        "bledy": bledy,
        "source_name": source_name,
        "source_path": source_path,
    }
    if import_record is not None:
        import_record.liczba_rekordow = dodane
        import_record.liczba_bledow = len(bledy)
        import_record.komunikat = " | ".join(bledy[:10]) if bledy else f"Zaimportowano {dodane} pozycji."
        import_record.save(update_fields=["liczba_rekordow", "liczba_bledow", "komunikat", "updated_at"])
    return summary


def import_koszty_excel_from_path(*, username: str, path: str):
    User = get_user_model()
    user = User.objects.get(username=username)
    src = Path(path)
    checksum = hashlib.sha256(src.read_bytes()).hexdigest() if src.exists() and src.is_file() else ""
    log_demon_event(
        modul=DemonLog.MODUL_FIRMA,
        poziom=DemonLog.POZIOM_INFO,
        wiadomosc="Rozpoczęto import pliku firmowego.",
        nazwa_pliku=src.name,
        sciezka=str(src),
        checksum_sha256=checksum,
    )
    import_record = ImportDemona.objects.create(
        user=user,
        nazwa_pliku=src.name,
        sciezka_zrodlowa=str(src),
        status=ImportDemona.STATUS_PRZETWARZANY,
        checksum_sha256=checksum,
    )
    try:
        with src.open("rb") as f:
            summary = import_koszty_excel_for_user(
                user=user,
                fileobj=f,
                source_name=src.name,
                source_path=str(src),
                import_record=import_record,
            )
        import_record.status = ImportDemona.STATUS_OK if not summary["bledy"] else ImportDemona.STATUS_CZESCIOWO
        import_record.save(update_fields=["status", "updated_at"])
        log_demon_event(
            modul=DemonLog.MODUL_FIRMA,
            poziom=DemonLog.POZIOM_WARNING if summary["bledy"] else DemonLog.POZIOM_INFO,
            wiadomosc=f"Import firmowy zakończony. Dodano {summary['dodane']} rekordów, błędów {len(summary['bledy'])}.",
            nazwa_pliku=src.name,
            sciezka=str(src),
            checksum_sha256=checksum,
        )
        return summary
    except Exception as exc:
        import_record.status = ImportDemona.STATUS_BLAD
        import_record.komunikat = str(exc)
        import_record.save(update_fields=["status", "komunikat", "updated_at"])
        log_demon_event(
            modul=DemonLog.MODUL_FIRMA,
            poziom=DemonLog.POZIOM_ERROR,
            wiadomosc=f"Import firmowy zakończony błędem: {exc}",
            nazwa_pliku=src.name,
            sciezka=str(src),
            checksum_sha256=checksum,
        )
        raise
